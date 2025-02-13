import pandas as pd
from openpyxl import load_workbook

# 🔹 Remplace par le chemin de ton fichier Excel contenant les Work Orders classés par type d’équipement
file_path = r"C:\Users\amouaoui\cernbox\WINDOWS\Desktop\Documents stage_CERN\Récuperation des données sous format excel\classifications des WO\Work_Orders_Classified.xlsx"

# Charger toutes les feuilles du fichier
sheets = pd.read_excel(file_path, sheet_name=None)

# 🔹 Nom des colonnes
equipment_column = "Equipement"
date_column = "Date Created"  # Remplace par le bon nom si différent

# 🔹 Nom du fichier de sortie
output_path = "Work_Orders_Grouped_By_EquipmentVf.xlsx"

# Créer un fichier Excel avec toutes les feuilles bien organisées
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    for sheet_name, df in sheets.items():  # Parcourir chaque feuille
        if equipment_column not in df.columns or date_column not in df.columns:
            print(f"⚠️ La feuille '{sheet_name}' ne contient pas les bonnes colonnes, elle est ignorée.")
            continue

        # Trier les Work Orders par Equipement et Date
        df = df.sort_values(by=[equipment_column, date_column])

        # 🔹 Créer un DataFrame vide avec les mêmes colonnes
        grouped_df = pd.DataFrame(columns=df.columns)

        # 🔹 Insérer une ligne vide correctement formatée
        separator = pd.DataFrame([{col: None for col in df.columns}])

        for equip, data in df.groupby(equipment_column):
            # Vérifier que le DataFrame `data` n'est pas vide avant de concaténer
            if not data.empty:
                grouped_df = pd.concat([grouped_df, data, separator], ignore_index=True)

        # Écriture dans une nouvelle feuille du fichier Excel
        grouped_df.to_excel(writer, sheet_name=sheet_name, index=False)

# 🔹 Ajuster automatiquement la largeur des colonnes
wb = load_workbook(output_path)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Lettre de la colonne (A, B, C, ...)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        sheet.column_dimensions[col_letter].width = max_length + 2  # Ajoute un peu d'espace pour plus de lisibilité

# Sauvegarder le fichier avec les colonnes ajustées
wb.save(output_path)

print(f"✅ Fichier enregistré : {output_path} avec les colonnes ajustées automatiquement !")
